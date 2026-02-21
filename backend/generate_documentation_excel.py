"""
Generate comprehensive Excel documentation for DVBC ERP System
Covers: Flows, Features, Permissions, Roles, Levels, Pages, Department Interconnectivity
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_sheet(ws, header_row=1):
    """Apply styling to worksheet"""
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

# =====================================================
# SHEET 1: ALL ROLES & PERMISSIONS
# =====================================================
ws_roles = wb.active
ws_roles.title = "Roles & Permissions"

roles_data = [
    ["Role ID", "Role Name", "Description", "Is System Role", "Can Delete", "Category", "Default Level"],
    ["admin", "Admin", "Full system access - all modules, all data, all actions", True, False, "System", "Leader"],
    ["manager", "Manager", "View/Download access, approve agreements, team management", True, False, "System", "Manager"],
    ["executive", "Executive", "Sales team - create leads, SOW, quotations", True, False, "Sales", "Executive"],
    ["consultant", "Consultant", "View SOW, update progress/status, task management", True, False, "Consulting", "Executive"],
    ["project_manager", "Project Manager", "Audit, approve, authorize SOW for client", True, False, "Consulting", "Manager"],
    ["principal_consultant", "Principal Consultant", "Freeze authority, SOW approval, senior consultant", True, False, "Consulting", "Leader"],
    ["lean_consultant", "Lean Consultant", "Junior consultant role - limited access", False, True, "Consulting", "Executive"],
    ["lead_consultant", "Lead Consultant", "Lead consultant with team oversight", False, True, "Consulting", "Manager"],
    ["senior_consultant", "Senior Consultant", "Senior consultant with advanced permissions", False, True, "Consulting", "Manager"],
    ["hr_executive", "HR Executive", "HR team member - no consulting data access", False, True, "HR", "Executive"],
    ["hr_manager", "HR Manager", "HR team manager - full HR access, limited consulting view", False, True, "HR", "Manager"],
    ["sales_manager", "Account Manager", "Handles client accounts and sales", False, True, "Sales", "Manager"],
    ["subject_matter_expert", "Subject Matter Expert", "Domain expert for consulting", False, True, "Consulting", "Leader"],
]

for row in roles_data:
    ws_roles.append(row)
style_sheet(ws_roles)
ws_roles.column_dimensions['A'].width = 20
ws_roles.column_dimensions['B'].width = 22
ws_roles.column_dimensions['C'].width = 50
ws_roles.column_dimensions['D'].width = 15
ws_roles.column_dimensions['E'].width = 12
ws_roles.column_dimensions['F'].width = 12
ws_roles.column_dimensions['G'].width = 15

# =====================================================
# SHEET 2: EMPLOYEE LEVELS & PERMISSIONS
# =====================================================
ws_levels = wb.create_sheet("Employee Levels")

levels_data = [
    ["Level", "Description", "can_view_team_data", "can_approve_requests", "can_manage_team", "can_view_reports", "can_edit_records", "can_delete_records", "can_export_data", "can_access_admin", "can_manage_permissions", "can_view_salary_data"],
    ["Executive", "Entry-level employees - Basic access to own data", True, False, False, False, False, False, False, False, False, False],
    ["Manager", "Mid-level employees - Team management capabilities", True, True, True, True, True, False, True, False, False, False],
    ["Leader", "Senior employees - Full department access", True, True, True, True, True, True, True, True, True, True],
]

for row in levels_data:
    ws_levels.append(row)
style_sheet(ws_levels)
ws_levels.column_dimensions['A'].width = 12
ws_levels.column_dimensions['B'].width = 45
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
    ws_levels.column_dimensions[col].width = 18

# =====================================================
# SHEET 3: DEPARTMENT-WISE FEATURES
# =====================================================
ws_dept = wb.create_sheet("Dept Features")

dept_data = [
    ["Department", "Feature", "Page/Route", "Start Point", "End Point", "Key Actions", "Connected Departments"],
    # SALES DEPARTMENT
    ["Sales", "Lead Management", "/leads", "New Lead Entry", "Lead Qualified/Lost", "Create, Edit, Score, Status Change, CSV Import", "Consulting (handover)"],
    ["Sales", "Meetings", "/meetings", "Meeting Created", "MOM Documented", "Schedule, Record MOM, Action Items, Follow-up", "Consulting"],
    ["Sales", "Pricing Plans", "/sales/pricing-plans", "Lead Qualified", "Pricing Finalized", "Create Plan, Add Components, Team Deployment", "Finance, Consulting"],
    ["Sales", "SOW Builder", "/sales/sow", "Pricing Approved", "SOW Submitted", "Create Scope Items, Add Categories, Submit", "Consulting, Admin"],
    ["Sales", "Quotations", "/sales/quotations", "SOW Approved", "Quotation Finalized", "Generate Quote, Negotiate, Finalize", "Finance"],
    ["Sales", "Proforma Invoice", "/sales/proforma", "Quote Finalized", "Invoice Generated", "Create Invoice, Version History, Proceed to Agreement", "Finance"],
    ["Sales", "Agreements", "/sales/agreements", "Proforma Accepted", "Agreement Signed", "Create Agreement, E-Signature, Download", "Admin, Consulting"],
    ["Sales", "Payment Verification", "/sales/payment-verification", "Agreement Signed", "Payment Verified", "Verify First Installment", "Finance, Consulting"],
    ["Sales", "Kickoff Request", "/kickoff-requests", "Payment Verified", "Kickoff Created", "Create Request, Submit to Consulting", "Consulting"],
    
    # HR DEPARTMENT
    ["HR", "Employee Management", "/employees", "New Hire", "Employee Active", "Create, Edit, Document Upload, Grant Access", "All Departments"],
    ["HR", "Onboarding", "/onboarding", "Employee Created", "Onboarding Complete", "5-Step Wizard, Documents, Bank Details", "Admin"],
    ["HR", "Attendance", "/attendance", "Daily Record", "Monthly Summary", "Mark Attendance, Work Location, Analytics", "Payroll"],
    ["HR", "Leave Management", "/leave-management", "Leave Request", "Leave Approved/Rejected", "Apply, Approve, Half-day Support, Withdraw", "Payroll"],
    ["HR", "Payroll", "/payroll", "Month Start", "Salary Slip Generated", "Calculate Salary, Deductions, Generate Slips", "Finance"],
    ["HR", "CTC Designer", "/ctc-designer", "CTC Submission", "CTC Approved", "Design CTC, Component Toggle, Admin Approval", "Finance, Admin"],
    ["HR", "Staffing Requests", "/hr/staffing-requests", "Request Created", "Request Fulfilled", "Budget, Skills, Timeline, Priority", "Consulting, Admin"],
    ["HR", "Travel Reimbursement", "/travel-reimbursement", "Claim Submitted", "Amount Reimbursed", "Distance Calc, Approval, Payroll Integration", "Finance"],
    ["HR", "Attendance Approvals", "/attendance-approvals", "Check-in Pending", "Check-in Approved", "Review Selfie, Location, Approve/Reject", "Mobile App"],
    ["HR", "Team Workload", "/hr/team-workload", "View Request", "Data Displayed", "View Consultant Utilization (Read-Only)", "Consulting"],
    
    # CONSULTING DEPARTMENT
    ["Consulting", "Projects", "/projects", "Kickoff Accepted", "Project Completed", "View, Manage, Status Update", "Sales, Finance"],
    ["Consulting", "Assign Team", "/consulting/assign-team/:id", "Project Created", "Team Assigned", "Add/Remove Consultants, Notifications", "HR"],
    ["Consulting", "My Projects", "/consultant/my-projects", "Assignment", "Project View", "View Assigned Projects, Tasks", "Self"],
    ["Consulting", "SOW List", "/consulting/sow-list", "Project Active", "SOW Viewed", "View Inherited SOW", "Sales"],
    ["Consulting", "Project Tasks", "/consulting/tasks/:id", "SOW Items", "Tasks Complete", "Create, Update Status, Link to SOW", "Self"],
    ["Consulting", "Timesheets", "/timesheets", "Week Start", "Week Submitted", "Log Hours, Project-wise, Submit", "HR, Finance"],
    ["Consulting", "Project Payments", "/projects/:id/payments", "Project Active", "All Payments Received", "View Schedule, Send Reminder, Record Payment", "Finance, Sales"],
    ["Consulting", "Change Requests", "/consulting/change-requests", "CR Created", "CR Applied", "Impact Analysis, Approval, Execute", "Sales, Admin"],
    
    # ADMIN DEPARTMENT
    ["Admin", "Dashboard", "/admin-dashboard", "Login", "KPIs Viewed", "View All Metrics, Quick Actions", "All Departments"],
    ["Admin", "Approvals Center", "/approvals-center", "Pending Items", "All Approved/Rejected", "CTC, Leaves, Expenses, Bank Changes", "All Departments"],
    ["Admin", "Permission Manager", "/permission-manager", "View Roles", "Permissions Updated", "Assign Roles, Edit Permissions", "All Departments"],
    ["Admin", "Role Management", "/role-management", "Role Request", "Role Assigned", "Create Role, Approval Workflow", "HR"],
    ["Admin", "Permission Dashboard", "/permission-dashboard", "View Levels", "Level Updated", "Edit Level Permissions, Assign Levels", "All Departments"],
    ["Admin", "Admin Masters", "/admin-masters", "View Config", "Config Updated", "Departments, Designations, SOW Scopes", "All Departments"],
    ["Admin", "Office Locations", "/office-locations", "View Locations", "Location Updated", "Geofencing Config, Add/Edit Offices", "HR, Mobile App"],
    ["Admin", "Security Audit", "/security-audit", "View Logs", "Audit Complete", "View Login Attempts, Password Changes", "All Departments"],
    
    # FINANCE (Implied through other modules)
    ["Finance", "Payment Tracking", "/payments", "Payment Due", "Payment Received", "View All Payments, Project-wise Summary", "Sales, Consulting"],
    ["Finance", "Expense Management", "/expenses", "Expense Submitted", "Expense Reimbursed", "Review, Approve, Link to Payroll", "HR"],
]

for row in dept_data:
    ws_dept.append(row)
style_sheet(ws_dept)
ws_dept.column_dimensions['A'].width = 12
ws_dept.column_dimensions['B'].width = 22
ws_dept.column_dimensions['C'].width = 30
ws_dept.column_dimensions['D'].width = 20
ws_dept.column_dimensions['E'].width = 20
ws_dept.column_dimensions['F'].width = 45
ws_dept.column_dimensions['G'].width = 25

# =====================================================
# SHEET 4: ROLE-WISE PAGE ACCESS
# =====================================================
ws_access = wb.create_sheet("Role Page Access")

# Header row
access_header = ["Page/Feature", "Route", "Admin", "Manager", "HR Manager", "HR Executive", "Project Manager", "Principal Consultant", "Consultant", "Account Manager", "Executive"]
ws_access.append(access_header)

access_data = [
    # My Workspace
    ["My Attendance", "/my-attendance", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["My Leaves", "/my-leaves", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["My Expenses", "/my-expenses", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["My Salary Slips", "/my-salary-slips", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["Bank Details Change", "/my-bank-details", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["Timesheets", "/timesheets", "✓", "✓", "—", "—", "✓", "✓", "✓", "—", "—"],
    
    # HR Section
    ["Employees", "/employees", "✓", "✓", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Attendance", "/attendance", "✓", "✓", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Leave Management", "/leave-management", "✓", "✓", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Payroll", "/payroll", "✓", "✓", "✓", "—", "—", "—", "—", "—", "—"],
    ["CTC Designer", "/ctc-designer", "✓", "—", "✓", "—", "—", "—", "—", "—", "—"],
    ["Team Workload", "/hr/team-workload", "✓", "✓", "✓", "—", "—", "—", "—", "—", "—"],
    ["Attendance Approvals", "/attendance-approvals", "✓", "✓", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Staffing Requests", "/hr/staffing-requests", "✓", "✓", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Onboarding", "/onboarding", "✓", "—", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Letter Management", "/letter-management", "✓", "—", "✓", "—", "—", "—", "—", "—", "—"],
    
    # Sales Section
    ["Leads", "/leads", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["Meetings (Sales)", "/meetings", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["Pricing Plans", "/sales/pricing-plans", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["SOW Builder", "/sales/sow", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["Quotations", "/sales/quotations", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["Proforma Invoice", "/sales/proforma", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["Agreements", "/sales/agreements", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["Payment Verification", "/sales/payment-verification", "✓", "✓", "—", "—", "—", "—", "—", "✓", "✓"],
    ["Kickoff Requests", "/kickoff-requests", "✓", "✓", "—", "—", "✓", "✓", "—", "✓", "✓"],
    
    # Consulting Section
    ["Projects", "/projects", "✓", "✓", "—", "—", "✓", "✓", "✓", "—", "—"],
    ["Assign Team", "/consulting/assign-team/:id", "✓", "✓", "—", "—", "✓", "✓", "—", "—", "—"],
    ["My Projects", "/consultant/my-projects", "✓", "✓", "—", "—", "✓", "✓", "✓", "—", "—"],
    ["Project Tasks", "/consulting/tasks/:id", "✓", "✓", "—", "—", "✓", "✓", "✓", "—", "—"],
    ["SOW List (Consulting)", "/consulting/sow-list", "✓", "✓", "—", "—", "✓", "✓", "✓", "—", "—"],
    ["Project Payments", "/projects/:id/payments", "✓", "✓", "—", "—", "✓", "✓", "✓", "—", "—"],
    ["Change Requests", "/consulting/change-requests", "✓", "✓", "—", "—", "✓", "✓", "✓", "—", "—"],
    
    # Admin Section
    ["Admin Dashboard", "/admin-dashboard", "✓", "✓", "—", "—", "—", "—", "—", "—", "—"],
    ["Approvals Center", "/approvals-center", "✓", "✓", "✓", "—", "—", "—", "—", "—", "—"],
    ["Permission Manager", "/permission-manager", "✓", "—", "—", "—", "—", "—", "—", "—", "—"],
    ["Role Management", "/role-management", "✓", "—", "—", "—", "—", "—", "—", "—", "—"],
    ["Permission Dashboard", "/permission-dashboard", "✓", "—", "—", "—", "—", "—", "—", "—", "—"],
    ["Admin Masters", "/admin-masters", "✓", "—", "—", "—", "—", "—", "—", "—", "—"],
    ["Security Audit", "/security-audit", "✓", "—", "—", "—", "—", "—", "—", "—", "—"],
    ["User Management", "/user-management", "✓", "—", "—", "—", "—", "—", "—", "—", "—"],
    
    # Reports & Analytics
    ["Reports", "/reports", "✓", "✓", "✓", "—", "✓", "✓", "—", "—", "—"],
    ["Performance Dashboard", "/performance-dashboard", "✓", "✓", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Employee Scorecard", "/employee-scorecard", "✓", "✓", "✓", "✓", "—", "—", "—", "—", "—"],
    ["Workflow Diagrams", "/workflow", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
]

for row in access_data:
    ws_access.append(row)
style_sheet(ws_access)
ws_access.column_dimensions['A'].width = 22
ws_access.column_dimensions['B'].width = 28
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws_access.column_dimensions[col].width = 14

# =====================================================
# SHEET 5: BUSINESS FLOWS (End-to-End)
# =====================================================
ws_flows = wb.create_sheet("Business Flows")

flows_data = [
    ["Flow Name", "Category", "Step #", "Step Name", "Actor/Role", "Action", "System Event", "Next Step Trigger"],
    
    # Lead to Delivery Flow
    ["Lead to Delivery", "Sales→Consulting", 1, "Lead Creation", "Executive/Account Manager", "Create new lead with contact details", "Lead Score Calculated", "Lead enters pipeline"],
    ["Lead to Delivery", "Sales→Consulting", 2, "Lead Qualification", "Executive/Account Manager", "Update status, add notes, schedule meetings", "Status Change Logged", "Lead marked 'Hot'"],
    ["Lead to Delivery", "Sales→Consulting", 3, "Pricing Plan", "Executive/Account Manager", "Create pricing with components, team deployment", "Plan Created", "Plan saved"],
    ["Lead to Delivery", "Sales→Consulting", 4, "SOW Creation", "Executive/Account Manager", "Build scope items, categories, deliverables", "SOW Version Created", "Submit for approval"],
    ["Lead to Delivery", "Sales→Consulting", 5, "SOW Approval", "Manager/Admin", "Review and approve SOW items", "Approval Logged", "SOW approved"],
    ["Lead to Delivery", "Sales→Consulting", 6, "Quotation", "Executive/Account Manager", "Generate quotation from SOW", "Quotation Created", "Quote finalized"],
    ["Lead to Delivery", "Sales→Consulting", 7, "Proforma Invoice", "Executive/Account Manager", "Create proforma invoice with payment terms", "Invoice Generated", "Proceed to agreement"],
    ["Lead to Delivery", "Sales→Consulting", 8, "Agreement", "Executive/Account Manager", "Create agreement with e-signature", "Agreement Created", "Client signs"],
    ["Lead to Delivery", "Sales→Consulting", 9, "Payment Verification", "Finance/Admin", "Verify first installment payment", "Payment Marked Received", "Payment verified"],
    ["Lead to Delivery", "Sales→Consulting", 10, "Kickoff Request", "Sales Team", "Create kickoff request with project details", "Request Created", "Submit to consulting"],
    ["Lead to Delivery", "Sales→Consulting", 11, "Kickoff Acceptance", "PM/Principal Consultant", "Accept kickoff request", "Project Created (status=active)", "Project starts"],
    ["Lead to Delivery", "Sales→Consulting", 12, "Team Assignment", "PM/Principal/Admin", "Assign consultants to project", "Notifications Sent", "Team notified"],
    ["Lead to Delivery", "Sales→Consulting", 13, "Project Execution", "Consultants", "Execute tasks, log timesheets", "Progress Updated", "Ongoing"],
    ["Lead to Delivery", "Sales→Consulting", 14, "Payment Collection", "Finance", "Track installments, send reminders", "Payments Logged", "Project complete"],
    
    # Employee Lifecycle
    ["Employee Lifecycle", "HR", 1, "Onboarding Start", "HR Executive", "Create employee record with personal details", "Employee ID Generated", "Proceed to employment"],
    ["Employee Lifecycle", "HR", 2, "Employment Details", "HR Executive", "Add department, designation, reporting manager", "Record Updated", "Proceed to documents"],
    ["Employee Lifecycle", "HR", 3, "Documents Upload", "HR Executive/Employee", "Upload Aadhaar, PAN, educational certificates", "Documents Stored", "Proceed to bank"],
    ["Employee Lifecycle", "HR", 4, "Bank Details", "HR Executive/Employee", "Add bank account with proof", "Bank Details Saved", "Complete onboarding"],
    ["Employee Lifecycle", "HR", 5, "Grant Portal Access", "HR Manager/Admin", "Create login credentials with temp password", "User Account Created", "Employee can login"],
    ["Employee Lifecycle", "HR", 6, "CTC Design", "HR Manager", "Design CTC structure with components", "CTC Submitted for Approval", "Admin reviews"],
    ["Employee Lifecycle", "HR", 7, "CTC Approval", "Admin", "Review and approve CTC", "CTC Activated", "Payroll ready"],
    ["Employee Lifecycle", "HR", 8, "Daily Operations", "Employee", "Attendance, leaves, expenses, timesheets", "Records Logged", "Ongoing"],
    ["Employee Lifecycle", "HR", 9, "Salary Processing", "HR Manager/Admin", "Generate salary slips", "Payroll Generated", "Monthly cycle"],
    
    # Leave Management
    ["Leave Management", "HR", 1, "Leave Application", "Employee", "Apply leave with dates, reason, type", "Leave Request Created", "Manager notified"],
    ["Leave Management", "HR", 2, "Manager Review", "Reporting Manager", "Review leave request", "Decision Made", "Approve/Reject"],
    ["Leave Management", "HR", 3, "Approval", "Manager", "Approve with optional comments", "Leave Approved", "Balance Updated"],
    ["Leave Management", "HR", 4, "Payroll Link", "System", "Auto-calculate leave days for payroll", "Deduction Calculated", "Salary slip reflects"],
    
    # Expense Reimbursement
    ["Expense Reimbursement", "HR→Finance", 1, "Expense Submission", "Employee", "Submit expense with receipt", "Expense Created", "Manager notified"],
    ["Expense Reimbursement", "HR→Finance", 2, "Manager Approval", "Reporting Manager", "Review and approve expense", "Expense Approved", "Finance notified"],
    ["Expense Reimbursement", "HR→Finance", 3, "Finance Processing", "Finance/Admin", "Process reimbursement", "Expense Paid", "Added to payroll"],
    
    # Bank Details Change
    ["Bank Details Change", "HR→Admin", 1, "Request Submission", "Employee", "Submit new bank details with proof", "Request Created (pending_hr)", "HR notified"],
    ["Bank Details Change", "HR→Admin", 2, "HR Review", "HR Manager", "Verify documents and approve", "Status: pending_admin", "Admin notified"],
    ["Bank Details Change", "HR→Admin", 3, "Admin Approval", "Admin", "Final approval", "Bank Details Updated", "Employee notified"],
    
    # Payment Reminders
    ["Payment Reminders", "Consulting→Finance", 1, "Check Eligibility", "System", "Check if within 7 days of due date", "Reminder Enabled", "Remind button active"],
    ["Payment Reminders", "Consulting→Finance", 2, "Send Reminder", "Consultant/PM", "Click remind button", "Notification Sent", "Finance, Sales, Admin notified"],
    ["Payment Reminders", "Consulting→Finance", 3, "Record Payment", "Consultant/PM", "Enter transaction ID after payment", "Payment Logged", "Incentive calculation triggered"],
]

for row in flows_data:
    ws_flows.append(row)
style_sheet(ws_flows)
ws_flows.column_dimensions['A'].width = 20
ws_flows.column_dimensions['B'].width = 18
ws_flows.column_dimensions['C'].width = 8
ws_flows.column_dimensions['D'].width = 22
ws_flows.column_dimensions['E'].width = 22
ws_flows.column_dimensions['F'].width = 40
ws_flows.column_dimensions['G'].width = 28
ws_flows.column_dimensions['H'].width = 22

# =====================================================
# SHEET 6: ALL PAGES INVENTORY
# =====================================================
ws_pages = wb.create_sheet("Pages Inventory")

pages_data = [
    ["Page Name", "Route", "File Path", "Department", "Functionality", "Key Components", "API Endpoints Used"],
    
    # Login & Auth
    ["Login", "/login", "Login.js", "System", "Main ERP login", "Password toggle, Portal links", "POST /api/auth/login"],
    ["HR Login", "/hr/login", "HRLogin.js", "HR", "HR Portal login", "Dedicated HR branding", "POST /api/auth/login"],
    ["Sales Login", "/sales/login", "SalesLogin.js", "Sales", "Sales Portal login", "Sales branding", "POST /api/auth/login"],
    ["Accept Offer", "/accept-offer/:token", "AcceptOfferPage.js", "HR", "Employee offer acceptance", "Company letterhead, Signature", "POST /api/letters/accept"],
    
    # Dashboards
    ["Admin Dashboard", "/admin-dashboard", "AdminDashboard.js", "Admin", "KPIs, Quick Actions, Charts", "Bento grid, Responsive", "GET /api/stats/dashboard"],
    ["Sales Dashboard", "/sales-dashboard", "SalesDashboardEnhanced.js", "Sales", "Pipeline, Targets, Leaderboard", "KPI cards, Charts", "GET /api/stats/sales-dashboard-enhanced"],
    ["HR Dashboard", "/hr-dashboard", "HRDashboard.js", "HR", "Employee stats, Attendance", "Charts, Quick links", "GET /api/stats/hr-dashboard"],
    ["Consulting Dashboard", "/consulting-dashboard", "ConsultingDashboard.js", "Consulting", "Project stats, Workload", "Charts, Efficiency", "GET /api/stats/consulting-dashboard"],
    ["HR Portal Dashboard", "/hr/dashboard", "HRPortalDashboard.js", "HR", "HR Portal home", "Quick actions, Stats", "GET /api/stats/hr-dashboard"],
    
    # HR Pages
    ["Employees", "/employees", "Employees.js", "HR", "Employee CRUD", "Table, Dialogs, Mobile toggle", "CRUD /api/employees"],
    ["Onboarding", "/onboarding", "HROnboarding.js", "HR", "5-step wizard", "Personal, Employment, Docs, Bank, Review", "POST /api/employees"],
    ["Attendance", "/attendance", "Attendance.js", "HR", "Company attendance", "Table, Mark dialog, Analytics", "CRUD /api/attendance"],
    ["My Attendance", "/my-attendance", "MyAttendance.js", "Self", "Personal attendance", "Check-in, Location", "GET /api/my/attendance"],
    ["Leave Management", "/leave-management", "LeaveManagement.js", "HR", "Company leaves", "Table, Approve/Reject", "CRUD /api/leave-requests"],
    ["My Leaves", "/my-leaves", "MyLeaves.js", "Self", "Personal leaves", "Apply, Withdraw, Half-day", "CRUD /api/leave-requests"],
    ["Payroll", "/payroll", "Payroll.js", "HR", "Salary processing", "Generate slips, Deductions", "/api/payroll/*"],
    ["My Salary Slips", "/my-salary-slips", "MySalarySlips.js", "Self", "View salary slips", "Download PDF", "/api/salary-slips/my"],
    ["Expenses", "/expenses", "Expenses.js", "HR", "Company expenses", "Approve, Reject", "CRUD /api/expenses"],
    ["My Expenses", "/my-expenses", "MyExpenses.js", "Self", "Personal expenses", "Submit, Receipt upload", "CRUD /api/expenses"],
    ["CTC Designer", "/ctc-designer", "CTCDesigner.js", "HR", "CTC structure", "Components toggle, Preview", "/api/ctc/*"],
    ["Team Workload", "/hr/team-workload", "HRTeamWorkload.js", "HR", "Consultant utilization", "Read-only view", "/api/consultants"],
    ["Attendance Approvals", "/attendance-approvals", "hr/HRAttendanceApprovals.js", "HR", "Mobile check-in approvals", "Selfie, Location review", "/api/hr/attendance-approvals"],
    ["Staffing Requests", "/hr/staffing-requests", "HRStaffingRequests.js", "HR", "Project staffing", "Form, Approval workflow", "/api/staffing-requests"],
    ["Letter Management", "/letter-management", "LetterManagement.js", "HR", "Offer/Appointment letters", "Templates, Send, Track", "/api/letters/*"],
    ["Performance Dashboard", "/performance-dashboard", "PerformanceDashboard.js", "HR", "Attendance analytics", "Charts, Work location", "/api/attendance/analytics"],
    ["Travel Reimbursement", "/travel-reimbursement", "TravelReimbursement.js", "HR", "Travel claims", "Distance calc, Approve", "/api/travel/*"],
    ["Bank Details Change", "/my-bank-details", "BankDetailsChangeRequest.js", "Self", "Bank change request", "IFSC verify, Proof upload", "/api/my/bank-change-request"],
    
    # Sales Pages
    ["Leads", "/leads", "Leads.js", "Sales", "Lead management", "Table/Card view, Search, CSV import", "CRUD /api/leads"],
    ["Sales Meetings", "/meetings", "SalesMeetings.js", "Sales", "Meeting management", "Schedule, MOM, Action items", "CRUD /api/meetings"],
    ["Pricing Plans", "/sales/pricing-plans", "sales-funnel/PricingPlanBuilder.js", "Sales", "Pricing builder", "Components, Team deployment", "/api/pricing-plans"],
    ["SOW Builder", "/sales/sow", "sales-funnel/SOWBuilder.js", "Sales", "SOW creation", "Scope items, Categories", "/api/sow/*"],
    ["Quotations", "/sales/quotations", "sales-funnel/Quotations.js", "Sales", "Quote generation", "Finalize, Negotiate", "/api/quotations"],
    ["Proforma Invoice", "/sales/proforma", "sales-funnel/ProformaInvoice.js", "Sales", "Invoice generation", "Version history", "/api/quotations"],
    ["Agreements", "/sales/agreements", "sales-funnel/Agreements.js", "Sales", "Contract management", "E-signature, Download", "/api/agreements"],
    ["Payment Verification", "/sales/payment-verification", "sales-funnel/PaymentVerification.js", "Sales", "First payment verify", "Confirm payment", "/api/payments"],
    ["Kickoff Requests", "/kickoff-requests", "KickoffRequests.js", "Sales/Consulting", "Project handover", "Create, Accept", "/api/kickoff-requests"],
    ["Clients", "/clients", "Clients.js", "Sales", "Client management", "Geo-coordinates", "CRUD /api/clients"],
    
    # Consulting Pages
    ["Projects", "/projects", "Projects.js", "Consulting", "Project list", "Status, Team", "/api/projects"],
    ["Assign Team", "/consulting/assign-team/:id", "consulting/AssignTeam.js", "Consulting", "Assign consultants", "Add/Remove members", "/api/projects/:id/team"],
    ["My Projects", "/consultant/my-projects", "consulting/MyProjects.js", "Consulting", "Assigned projects", "View own projects", "/api/consultant/my-projects"],
    ["Project Tasks", "/consulting/tasks/:id", "consulting/ConsultingProjectTasks.js", "Consulting", "Task management", "Create, Status update", "/api/projects/:id/tasks"],
    ["SOW List", "/consulting/sow-list", "consulting/ConsultingSOWList.js", "Consulting", "Inherited SOW", "View, Track progress", "/api/enhanced-sow/*"],
    ["Project Payments", "/payments", "ProjectPayments.js", "Consulting", "All payments", "Summary, Upcoming", "/api/project-payments/*"],
    ["Payment Details", "/projects/:id/payments", "ProjectPaymentDetails.js", "Consulting", "Project payment detail", "Schedule, Remind, Record", "/api/project-payments/*"],
    ["Change Requests", "/consulting/change-requests", "consulting/SOWChangeRequests.js", "Consulting", "SOW changes", "Create, Approve, Apply", "/api/sow-change-requests"],
    ["Timesheets", "/timesheets", "Timesheets.js", "Consulting", "Time logging", "Weekly entry, Submit", "/api/timesheets"],
    ["Consulting Meetings", "/consulting-meetings", "ConsultingMeetings.js", "Consulting", "Project meetings", "Schedule, Tracking", "/api/consulting-meetings/*"],
    
    # Admin Pages
    ["Approvals Center", "/approvals-center", "ApprovalsCenter.js", "Admin", "All approvals", "CTC, Leaves, Expenses, Bank", "/api/*/pending-approvals"],
    ["Permission Manager", "/permission-manager", "PermissionManager.js", "Admin", "Role permissions", "Assign, Edit", "/api/roles/*"],
    ["Role Management", "/role-management", "RoleManagement.js", "Admin", "Role requests", "Create, Approve workflow", "/api/role-management/*"],
    ["Permission Dashboard", "/permission-dashboard", "PermissionDashboard.js", "Admin", "Level permissions", "View/Edit by level", "/api/role-management/*"],
    ["Admin Masters", "/admin-masters", "AdminMasters.js", "Admin", "System config", "Departments, Designations, SOW Scopes", "/api/masters/*"],
    ["User Management", "/user-management", "UserManagement.js", "Admin", "User accounts", "Create, Deactivate", "/api/users/*"],
    ["Security Audit", "/security-audit", "SecurityAuditLog.js", "Admin", "Audit logs", "Login attempts, Password changes", "/api/security-logs"],
    ["Office Locations", "/office-locations", "OfficeLocationsSettings.js", "Admin", "Geofencing", "Add/Edit offices", "/api/office-locations"],
    ["Employee Scorecard", "/employee-scorecard", "EmployeeScorecard.js", "Admin/HR", "Employee overview", "Stats, Timeline, Linked records", "/api/employees/*"],
    
    # Mobile & Utility
    ["Mobile App", "/mobile", "EmployeeMobileApp.js", "Self", "Mobile attendance", "Selfie, GPS, Travel claim", "/api/my/*"],
    ["Mobile App Download", "/mobile-app", "MobileAppDownload.js", "Info", "App installation guide", "QR code, Instructions", "None"],
    ["Workflow Diagrams", "/workflow", "WorkflowPage.js", "Info", "Process visualization", "Animated steps", "None"],
    ["Tutorials", "/tutorials", "OnboardingTutorial.js", "Info", "How-to guides", "Step tracking", "None"],
]

for row in pages_data:
    ws_pages.append(row)
style_sheet(ws_pages)
ws_pages.column_dimensions['A'].width = 22
ws_pages.column_dimensions['B'].width = 28
ws_pages.column_dimensions['C'].width = 35
ws_pages.column_dimensions['D'].width = 12
ws_pages.column_dimensions['E'].width = 25
ws_pages.column_dimensions['F'].width = 32
ws_pages.column_dimensions['G'].width = 32

# =====================================================
# SHEET 7: DUPLICATE/UNIFIED FUNCTIONS
# =====================================================
ws_dup = wb.create_sheet("Duplicate Functions")

dup_data = [
    ["Function/Feature", "Location 1", "Location 2", "Status", "Recommended Action", "Notes"],
    ["Login", "Login.js", "HRLogin.js, SalesLogin.js", "Intentional Split", "Keep as-is", "Different portals need different branding"],
    ["Leave Application", "MyLeaves.js", "LeaveManagement.js (removed)", "FIXED", "Done", "Apply Leave button removed from HR page"],
    ["Attendance Recording", "Attendance.js", "MyAttendance.js, EmployeeMobileApp.js", "Intentional Split", "Keep as-is", "HR marks for others, Self for own, Mobile for GPS"],
    ["Dashboard Stats", "AdminDashboard.js", "SalesDashboardEnhanced.js, HRDashboard.js", "Intentional Split", "Keep as-is", "Different KPIs per role"],
    ["SOW Builder", "SOWBuilder.js", "ConsultingSOWList.js", "Different Purpose", "Keep as-is", "Sales creates, Consulting views inherited"],
    ["Meetings", "SalesMeetings.js", "ConsultingMeetings.js", "Different Purpose", "Keep as-is", "Sales meetings vs Project meetings"],
    ["Project Payments", "ProjectPayments.js", "ProjectPaymentDetails.js", "Related", "Keep as-is", "List vs Detail view"],
    ["Bank Details", "HROnboarding.js (Step 4)", "BankDetailsChangeRequest.js", "Different Purpose", "Keep as-is", "Initial setup vs Change request"],
    ["CTC Design", "CTCDesigner.js", "Payroll.js", "Linked", "Keep as-is", "CTC feeds into payroll calculations"],
    ["Staffing Requests", "HRStaffingRequests.js", "KickoffRequests.js", "Different Purpose", "Keep as-is", "HR staffing vs Sales kickoff"],
    ["User Profile", "UserProfile.js", "Employee view in Employees.js", "Related", "Keep as-is", "Self-edit vs HR view"],
]

for row in dup_data:
    ws_dup.append(row)
style_sheet(ws_dup)
ws_dup.column_dimensions['A'].width = 22
ws_dup.column_dimensions['B'].width = 25
ws_dup.column_dimensions['C'].width = 35
ws_dup.column_dimensions['D'].width = 18
ws_dup.column_dimensions['E'].width = 18
ws_dup.column_dimensions['F'].width = 45

# =====================================================
# SHEET 8: DEPARTMENT INTERCONNECTIVITY
# =====================================================
ws_connect = wb.create_sheet("Dept Interconnectivity")

connect_data = [
    ["From Dept", "To Dept", "Trigger Event", "Data Shared", "Notification", "Automated Action"],
    ["Sales", "Consulting", "Kickoff Request Accepted", "Project details, SOW, Pricing Plan", "PM/Consultants notified", "Project created with status=active"],
    ["Sales", "Finance", "Payment Verified", "Payment amount, Transaction ID", "Finance notified", "Payment logged"],
    ["Sales", "Admin", "Agreement Signed", "Agreement document", "Admin notified", "Approval logged"],
    ["HR", "All", "Employee Created", "Employee ID, Name, Role", "Manager notified", "Access pending"],
    ["HR", "Payroll", "Leave Approved", "Leave days, Type", "Auto-calculated", "Deduction in salary slip"],
    ["HR", "Finance", "Expense Approved", "Amount, Receipt", "Finance notified", "Reimbursement queued"],
    ["HR", "Admin", "CTC Submitted", "Salary breakdown", "Admin notified", "Pending approval"],
    ["HR", "Admin", "Bank Change Request", "New bank details, Proof", "Admin notified (after HR approval)", "2-step approval"],
    ["Consulting", "Finance", "Payment Reminder Sent", "Project, Installment, Amount", "Finance, Sales, Admin notified", "Reminder logged"],
    ["Consulting", "Finance", "Payment Recorded", "Transaction ID, Amount", "Finance, Sales notified", "Payment marked received"],
    ["Consulting", "HR", "Consultant Assigned", "Project, Role", "Consultant & Manager notified", "Workload updated"],
    ["Consulting", "Sales", "SOW Change Request", "CR details", "Sales notified", "Pending review"],
    ["Admin", "All", "Role Assigned", "New role, Permissions", "Employee notified", "Access updated"],
    ["Admin", "HR", "CTC Approved", "Final CTC structure", "HR & Employee notified", "Payroll config updated"],
    ["Admin", "HR", "Bank Change Approved", "Bank details", "Employee notified", "Bank details updated"],
    ["Finance", "HR", "Salary Generated", "Payslip", "Employee notified", "Slip available for download"],
    ["Mobile App", "HR", "Check-in (Unknown Location)", "Selfie, GPS, Justification", "HR notified", "Pending approval"],
]

for row in connect_data:
    ws_connect.append(row)
style_sheet(ws_connect)
ws_connect.column_dimensions['A'].width = 12
ws_connect.column_dimensions['B'].width = 12
ws_connect.column_dimensions['C'].width = 28
ws_connect.column_dimensions['D'].width = 35
ws_connect.column_dimensions['E'].width = 28
ws_connect.column_dimensions['F'].width = 35

# =====================================================
# SHEET 9: RECENT SESSIONS (Last 2 Days)
# =====================================================
ws_recent = wb.create_sheet("Recent Sessions")

recent_data = [
    ["Date", "Session", "Major Features/Fixes", "Department Impact", "Files Modified", "Status"],
    ["Feb 18", "Session 14", "Payment Reminders & Record Payments", "Consulting, Finance", "project_payments.py, ProjectPaymentDetails.js", "Complete"],
    ["Feb 18", "Session 14", "Role-Based Payment Visibility", "Consulting, Finance", "project_payments.py, ProjectPayments.js, ProjectPaymentDetails.js", "Complete"],
    ["Feb 18", "Session 14", "Project Status Auto-Active on Kickoff", "Consulting", "kickoff.py", "Complete"],
    ["Feb 18", "Session 14", "Reset Temp Password Endpoint", "HR, Admin", "employees.py", "Complete"],
    ["Feb 18", "Session 14", "Remove Amounts Hidden UI", "Consulting", "ProjectPaymentDetails.js", "Complete"],
    ["Feb 18", "Session 13", "Employee Levels System (Executive/Manager/Leader)", "All", "role_management.py, RoleManagement.js, PermissionDashboard.js", "Complete"],
    ["Feb 18", "Session 13", "Level-Based Permissions (10 permissions)", "All", "role_management.py, PermissionContext.js, Layout.js", "Complete"],
    ["Feb 18", "Session 13", "Role Request Approval Workflow", "HR, Admin", "role_management.py, RoleManagement.js", "Complete"],
    ["Feb 18", "Session 13", "Offer & Appointment Letter System", "HR", "letters.py, LetterManagement.js, AcceptOfferPage.js", "Complete"],
    ["Feb 18", "Session 12", "Server.py Refactoring Phase 1-3", "Backend", "stats.py, auth.py, leads.py, etc.", "Complete"],
    ["Feb 18", "Session 12", "Stats Router Complete Migration", "All Dashboards", "stats.py, server.py", "Complete"],
    ["Feb 18", "Session 12", "SOW Scope Builder in Admin Masters", "Admin", "AdminMasters.js", "Complete"],
    ["Feb 17", "Session 11", "PWA Install Prompt", "All", "PWAInstallPrompt.js, manifest.json, service-worker.js", "Complete"],
    ["Feb 17", "Session 11", "Bank Details Change Workflow", "HR, Admin, Self", "BankDetailsChangeRequest.js, server.py", "Complete"],
    ["Feb 17", "Session 11", "Medium-Priority Workflows Added", "All", "WorkflowPage.js", "Complete"],
    ["Feb 17", "Session 11", "Mobile Responsive Layouts", "All", "Layout.js, HRLayout.js, SalesLayout.js, AdminDashboard.js", "Complete"],
    ["Feb 17", "Session 10", "CTC Designer with Configurable Components", "HR, Admin", "CTCDesigner.js, ctc.py", "Complete"],
    ["Feb 17", "Session 10", "CTC Approval Workflow", "HR, Admin", "ctc.py, ApprovalsCenter.js", "Complete"],
]

for row in recent_data:
    ws_recent.append(row)
style_sheet(ws_recent)
ws_recent.column_dimensions['A'].width = 10
ws_recent.column_dimensions['B'].width = 12
ws_recent.column_dimensions['C'].width = 42
ws_recent.column_dimensions['D'].width = 22
ws_recent.column_dimensions['E'].width = 50
ws_recent.column_dimensions['F'].width = 12

# =====================================================
# SHEET 10: API ENDPOINTS SUMMARY
# =====================================================
ws_api = wb.create_sheet("API Endpoints")

api_data = [
    ["Router/Module", "Method", "Endpoint", "Description", "Roles Allowed", "Department"],
    # Auth
    ["auth", "POST", "/api/auth/login", "User login", "All", "System"],
    ["auth", "POST", "/api/auth/register", "User registration", "Admin", "System"],
    ["auth", "POST", "/api/auth/change-password", "Change password", "All authenticated", "System"],
    ["auth", "POST", "/api/auth/google", "Google OAuth", "All", "System"],
    
    # Stats
    ["stats", "GET", "/api/stats/dashboard", "Main dashboard stats", "Admin, Manager", "Admin"],
    ["stats", "GET", "/api/stats/sales-dashboard", "Sales pipeline stats", "Sales, Admin", "Sales"],
    ["stats", "GET", "/api/stats/sales-dashboard-enhanced", "Enhanced sales metrics", "Sales, Admin", "Sales"],
    ["stats", "GET", "/api/stats/hr-dashboard", "HR employee/attendance stats", "HR, Admin", "HR"],
    ["stats", "GET", "/api/stats/consulting-dashboard", "Consulting delivery stats", "Consulting, Admin", "Consulting"],
    
    # Employees
    ["employees", "GET", "/api/employees", "List employees", "HR, Admin", "HR"],
    ["employees", "POST", "/api/employees", "Create employee", "HR, Admin", "HR"],
    ["employees", "PATCH", "/api/employees/{id}", "Update employee", "HR, Admin", "HR"],
    ["employees", "POST", "/api/employees/{id}/grant-access", "Grant portal access", "HR, Admin", "HR"],
    ["employees", "POST", "/api/employees/{id}/reset-temp-password", "Reset to temp password", "HR Manager, Admin", "HR"],
    
    # Role Management
    ["role_management", "GET", "/api/role-management/levels", "List employee levels", "All authenticated", "Admin"],
    ["role_management", "GET", "/api/role-management/level-permissions", "Get level permissions", "All authenticated", "Admin"],
    ["role_management", "PUT", "/api/role-management/level-permissions", "Update level permissions", "Admin", "Admin"],
    ["role_management", "POST", "/api/role-management/role-requests", "Create role request", "HR", "Admin"],
    ["role_management", "POST", "/api/role-management/role-requests/{id}/approve", "Approve role request", "Admin", "Admin"],
    
    # Project Payments
    ["project_payments", "GET", "/api/project-payments/project/{id}", "Project payment details", "Consulting, Admin", "Consulting"],
    ["project_payments", "GET", "/api/project-payments/my-payments", "User's project payments", "All authenticated", "Consulting"],
    ["project_payments", "GET", "/api/project-payments/upcoming", "Upcoming payments", "Admin, PM, Principal", "Finance"],
    ["project_payments", "POST", "/api/project-payments/send-reminder", "Send payment reminder", "Consulting team", "Consulting"],
    ["project_payments", "POST", "/api/project-payments/record-payment", "Record payment transaction", "Consulting team", "Consulting"],
    
    # Kickoff
    ["kickoff", "GET", "/api/kickoff-requests", "List kickoff requests", "Sales, Consulting, Admin", "Sales"],
    ["kickoff", "POST", "/api/kickoff-requests", "Create kickoff request", "Sales", "Sales"],
    ["kickoff", "PUT", "/api/kickoff/accept/{id}", "Accept kickoff (creates project)", "PM, Principal, Admin", "Consulting"],
    
    # CTC
    ["ctc", "POST", "/api/ctc/calculate-preview", "Preview CTC breakdown", "HR, Admin", "HR"],
    ["ctc", "POST", "/api/ctc/design", "Submit CTC for approval", "HR", "HR"],
    ["ctc", "GET", "/api/ctc/pending-approvals", "List pending CTC approvals", "Admin", "Admin"],
    ["ctc", "POST", "/api/ctc/{id}/approve", "Approve CTC", "Admin", "Admin"],
    
    # Attendance
    ["attendance", "GET", "/api/attendance", "List attendance records", "HR, Admin", "HR"],
    ["attendance", "POST", "/api/attendance", "Mark attendance", "HR, Admin", "HR"],
    ["attendance", "GET", "/api/attendance/analytics", "Attendance analytics", "HR, Admin", "HR"],
    ["attendance", "POST", "/api/my/check-in", "Self check-in (mobile)", "All authenticated", "Self"],
    
    # Leaves
    ["server", "GET", "/api/leave-requests", "List leave requests", "All authenticated", "HR"],
    ["server", "POST", "/api/leave-requests", "Apply for leave", "All authenticated", "Self"],
    ["server", "POST", "/api/leave-requests/{id}/approve", "Approve leave", "Manager, HR, Admin", "HR"],
    ["server", "POST", "/api/leave-requests/{id}/withdraw", "Withdraw leave request", "Requester only", "Self"],
    
    # Letters
    ["letters", "POST", "/api/letters/offer-letters", "Create offer letter", "HR, Admin", "HR"],
    ["letters", "POST", "/api/letters/offer-letters/accept", "Accept offer letter", "Public (with token)", "HR"],
    ["letters", "GET", "/api/letters/view/offer/{token}", "View offer letter", "Public (with token)", "HR"],
]

for row in api_data:
    ws_api.append(row)
style_sheet(ws_api)
ws_api.column_dimensions['A'].width = 18
ws_api.column_dimensions['B'].width = 8
ws_api.column_dimensions['C'].width = 45
ws_api.column_dimensions['D'].width = 35
ws_api.column_dimensions['E'].width = 25
ws_api.column_dimensions['F'].width = 12

# =====================================================
# Save workbook
# =====================================================
output_path = "/app/DVBC_ERP_Documentation.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
print(f"Sheets: {wb.sheetnames}")

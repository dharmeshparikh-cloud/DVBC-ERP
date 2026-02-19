"""
Add Employee Linkage sheet to existing Excel documentation
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load existing workbook
wb = load_workbook("/app/DVBC_ERP_Documentation.xlsx")

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_sheet(ws, header_row=1):
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

# =====================================================
# SHEET: Employee Data Linkage
# =====================================================
ws_link = wb.create_sheet("Employee Linkage", 0)  # Insert at beginning

linkage_data = [
    ["Module", "Collection", "Link Field", "How It Connects", "Example Query", "When Created"],
    
    # Section: Identity
    ["IDENTITY", "employees", "id", "Primary Key - used everywhere internally", "employees.find_one({id: X})", "HR Onboarding Step 1"],
    ["IDENTITY", "employees", "employee_id", "Display code (EMP001)", "Show in UI, reports", "Auto-generated on create"],
    ["IDENTITY", "employees", "user_id", "Links to users collection for login", "employees.user_id = users.id", "When portal access granted"],
    ["IDENTITY", "users", "id", "User's login account ID", "Used for notifications, assignments", "Grant Access action"],
    ["IDENTITY", "users", "employee_id", "Back-reference to employee code", "Cross-lookup", "Grant Access action"],
    
    # Section: HR Functions
    ["HR", "attendance", "employee_id", "Daily attendance records", "attendance.find({employee_id: X})", "Daily mark attendance"],
    ["HR", "leave_requests", "employee_id", "Leave applications", "leave_requests.find({employee_id: X})", "Employee applies leave"],
    ["HR", "salary_slips", "employee_id", "Monthly salary slips", "salary_slips.find({employee_id: X})", "Payroll generation"],
    ["HR", "ctc_structures", "employee_id", "CTC design records", "ctc_structures.find({employee_id: X})", "CTC Designer submit"],
    ["HR", "expenses", "employee_id", "Expense claims", "expenses.find({employee_id: X})", "Employee submits expense"],
    ["HR", "bank_change_requests", "employee_id", "Bank detail change requests", "bank_change_requests.find({employee_id: X})", "Employee requests change"],
    ["HR", "employee_documents", "employee_id", "Uploaded documents (Aadhaar, PAN)", "employee_documents.find({employee_id: X})", "Onboarding Step 3"],
    ["HR", "timesheets", "employee_id", "Weekly time entries", "timesheets.find({employee_id: X})", "Weekly timesheet submit"],
    ["HR", "travel_claims", "employee_id", "Travel reimbursement claims", "travel_claims.find({employee_id: X})", "Travel claim submit"],
    
    # Section: Sales Functions
    ["SALES", "leads", "created_by", "Lead creator (user.id)", "leads.find({created_by: user.id})", "Lead creation"],
    ["SALES", "leads", "assigned_to", "Lead owner (user.id)", "leads.find({assigned_to: user.id})", "Lead assignment"],
    ["SALES", "meetings", "created_by", "Meeting scheduler", "meetings.find({created_by: user.id})", "Meeting created"],
    ["SALES", "pricing_plans", "created_by", "Plan author", "pricing_plans.find({created_by: user.id})", "Pricing plan saved"],
    ["SALES", "sow", "created_by", "SOW author", "sow.find({created_by: user.id})", "SOW creation"],
    ["SALES", "quotations", "created_by", "Quote generator", "quotations.find({created_by: user.id})", "Quote generated"],
    ["SALES", "agreements", "created_by", "Agreement creator", "agreements.find({created_by: user.id})", "Agreement drafted"],
    ["SALES", "agreements", "signed_by_user_id", "Who signed", "For audit trail", "Agreement signed"],
    ["SALES", "kickoff_requests", "created_by", "Handoff initiator", "kickoff_requests.find({created_by: user.id})", "Kickoff submitted"],
    
    # Section: Consulting Functions
    ["CONSULTING", "project_assignments", "consultant_id", "Assigned consultant (user.id)", "project_assignments.find({consultant_id: user.id})", "Team assignment"],
    ["CONSULTING", "project_tasks", "assigned_to", "Task owner (user.id)", "project_tasks.find({assigned_to: user.id})", "Task assignment"],
    ["CONSULTING", "sow_items", "assigned_consultant_id", "SOW item assignee", "sow_items.find({assigned_consultant_id: user.id})", "SOW item assign"],
    ["CONSULTING", "consulting_meetings", "participants", "Meeting attendee (in array)", "consulting_meetings.find({participants: user.id})", "Meeting invite"],
    ["CONSULTING", "payment_reminders", "sent_by", "Who sent reminder", "For audit", "Remind button click"],
    ["CONSULTING", "installment_payments", "recorded_by", "Who recorded payment", "For audit", "Record payment action"],
    
    # Section: Admin/System
    ["SYSTEM", "notifications", "user_id", "Notification recipient", "notifications.find({user_id: user.id})", "Any system event"],
    ["SYSTEM", "approval_requests", "submitted_by", "Request submitter", "approval_requests.find({submitted_by: user.id})", "Any approval submit"],
    ["SYSTEM", "approval_requests", "approved_by", "Request approver", "approval_requests.find({approved_by: user.id})", "Approval action"],
    ["SYSTEM", "security_logs", "user_id", "Audit subject", "security_logs.find({user_id: user.id})", "Login, password change"],
    ["SYSTEM", "role_requests", "employee_id", "Role change target", "role_requests.find({employee_id: X})", "Role request submit"],
    ["SYSTEM", "role_requests", "submitted_by", "Who requested", "role_requests.find({submitted_by: user.id})", "HR submits"],
    
    # Section: Cross-Department Links
    ["HANDOFF", "projects", "created_from_kickoff", "Links project to kickoff request", "projects.find({created_from_kickoff: kickoff.id})", "Kickoff acceptance"],
    ["HANDOFF", "projects", "agreement_id", "Links project to agreement", "projects.find({agreement_id: agreement.id})", "Project creation"],
    ["HANDOFF", "sow (inherited)", "project_id", "SOW copied to project", "sow.find({project_id: project.id})", "Kickoff acceptance"],
]

for row in linkage_data:
    ws_link.append(row)

style_sheet(ws_link)
ws_link.column_dimensions['A'].width = 14
ws_link.column_dimensions['B'].width = 22
ws_link.column_dimensions['C'].width = 22
ws_link.column_dimensions['D'].width = 40
ws_link.column_dimensions['E'].width = 45
ws_link.column_dimensions['F'].width = 25

# Color code the Module column
module_colors = {
    "IDENTITY": "1F4E79",
    "HR": "2E7D32",
    "SALES": "F57C00",
    "CONSULTING": "7B1FA2",
    "SYSTEM": "616161",
    "HANDOFF": "C62828"
}

for row_idx in range(2, ws_link.max_row + 1):
    cell = ws_link.cell(row=row_idx, column=1)
    module = cell.value
    if module in module_colors:
        cell.fill = PatternFill(start_color=module_colors[module], end_color=module_colors[module], fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

# =====================================================
# SHEET: Employee Journey Timeline
# =====================================================
ws_journey = wb.create_sheet("Employee Journey", 1)

journey_data = [
    ["Step #", "Stage", "Action", "Who Does It", "Data Created", "Collections Updated", "Notifications Sent"],
    [1, "Hire", "Create Employee Record", "HR Executive", "Personal details, department, designation", "employees", "None"],
    [2, "Onboard", "Employment Details", "HR Executive", "DOJ, reporting manager, role", "employees", "Manager notified"],
    [3, "Onboard", "Upload Documents", "HR/Employee", "Aadhaar, PAN, certificates", "employees, employee_documents", "None"],
    [4, "Onboard", "Bank Details", "HR/Employee", "Account number, IFSC, proof", "employees", "None"],
    [5, "Access", "Grant Portal Access", "HR Manager/Admin", "Login credentials, temp password", "users, employees.user_id", "Employee gets email"],
    [6, "Setup", "Design CTC", "HR Manager", "Salary components, breakdown", "ctc_structures (pending)", "Admin notified"],
    [7, "Setup", "Approve CTC", "Admin", "CTC activation", "ctc_structures (active), employees.salary", "HR & Employee notified"],
    [8, "Daily", "Mark Attendance", "Self/HR", "Daily record", "attendance", "None"],
    [9, "Daily", "Apply Leave", "Self", "Leave request", "leave_requests, approval_requests", "Manager notified"],
    [10, "Daily", "Submit Expense", "Self", "Expense claim", "expenses, approval_requests", "Manager notified"],
    [11, "Daily", "Log Timesheet", "Self", "Weekly hours", "timesheets", "PM notified on submit"],
    [12, "Project", "Get Assigned to Project", "PM/Admin", "Team membership", "project_assignments", "Employee & Manager notified"],
    [13, "Project", "Work on Tasks", "Self", "Task updates", "project_tasks", "PM notified"],
    [14, "Project", "View SOW Items", "Self", "Read-only", "None (read)", "None"],
    [15, "Monthly", "Salary Generated", "HR/System", "Payslip", "salary_slips", "Employee notified"],
    [16, "Change", "Request Bank Change", "Self", "New bank details", "bank_change_requests", "HR notified"],
    [17, "Change", "Request Role Change", "HR", "Role request", "role_requests", "Admin notified"],
    [18, "Exit", "Terminate Employee", "HR/Admin", "Status update", "employees.status = terminated", "System access revoked"],
]

for row in journey_data:
    ws_journey.append(row)

style_sheet(ws_journey)
ws_journey.column_dimensions['A'].width = 8
ws_journey.column_dimensions['B'].width = 12
ws_journey.column_dimensions['C'].width = 25
ws_journey.column_dimensions['D'].width = 18
ws_journey.column_dimensions['E'].width = 35
ws_journey.column_dimensions['F'].width = 35
ws_journey.column_dimensions['G'].width = 25

# Save workbook
wb.save("/app/DVBC_ERP_Documentation.xlsx")
print("Updated Excel with Employee Linkage and Journey sheets")
print(f"Total sheets: {wb.sheetnames}")
